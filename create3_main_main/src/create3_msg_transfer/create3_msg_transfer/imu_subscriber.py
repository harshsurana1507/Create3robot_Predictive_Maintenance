import pandas as pd
import rclpy
from rclpy.node import Node
from sensor_msgs.msg import Imu
from rclpy.qos import QoSProfile, ReliabilityPolicy, LivelinessPolicy, DurabilityPolicy
from openpyxl import Workbook
from datetime import datetime




class ImuSubscriber(Node):
    def __init__(self):
        super().__init__('imu_subscriber')

        topic_name = '/robot01/imu'
        #change /robot01 with the name of the robot kept
        qos_profile = QoSProfile(
            reliability=ReliabilityPolicy.BEST_EFFORT,
            liveliness=LivelinessPolicy.AUTOMATIC,
            durability=DurabilityPolicy.VOLATILE,
            depth=1
        )

        self.subscription = self.create_subscription(
            Imu,
            topic_name,
            self.imu_callback,
            qos_profile
            )
        self.subscription  # prevent unused variable warning
        self.workbook = Workbook()
        self.orientation_worksheet = self.workbook.active
        # 3 seperate sheets are created for orientation, angular velocity and acceleration data for clarity
        self.orientation_worksheet.title = 'Orientation data '
        self.orientation_worksheet.append(["Timestamp", "imu_orient_X", "imu_orient_Y","imu_orient_Z","imu_orient_W"])
        #above line is the heading for orientation data sheet


        self.velocity_worksheet = self.workbook.create_sheet(title= 'Angular Velocity data ')
        self.velocity_worksheet.append(["Timestamp", "imu_angvel_X", "imu_angvel_Y","imu_angvel_Z"])
        #above line is the heading for angular velocity data sheet

        self.acceleration_worksheet = self.workbook.create_sheet(title= 'Acceleration data ')
        self.acceleration_worksheet.append(["Timestamp", "imu_acc_X", "imu_acc_Y","imu_acc_Z"])
        #above line is the heading for acceleration data sheet


        #all the data which will be added in the excel sheets, are first added in these list, then once the node is terminated
        # all the data from the list will be then saved in the excel file.
        self.orientation_data = []
        self.velocity_data = []
        self.acceleration_data = []

        self.timer = self.create_timer(1.0, self.process_buffered_messages)
        # in gap of 1 seconds the function process_buffered_messages will be called, so that the latest message from the topic which
        # was stored in the lists self.message_buffer can be extracted and stored in the data list.


        self.message_buffer = [] # a list is created so that all the readings from the topic can be stored in it.



    def imu_callback(self, msg):
        self.message_buffer.append(msg)

    
    def process_buffered_messages(self):
        if not self.message_buffer:
            return

        # Process the latest message from the buffer
        latest_msg = self.message_buffer[-1]

        #self.get_logger().info(str(msg))
        self.get_logger().info(f"IMU (Inertial Measurement Unit-new): \n - Orientation data --> x: {latest_msg.orientation.x} y: {latest_msg.orientation.y} z: {latest_msg.orientation.z} w:{latest_msg.orientation.w}\n - Angular Velocity data-->  x:{latest_msg.angular_velocity.x} y: {latest_msg.angular_velocity.y} z: {latest_msg.angular_velocity.z} \n - Linear acceleration data -->x: {latest_msg.linear_acceleration.x} y: {latest_msg.linear_acceleration.y} z: {latest_msg.linear_acceleration.z}")
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.orientation_data.append([current_time,latest_msg.orientation.x,latest_msg.orientation.y,latest_msg.orientation.z,latest_msg.orientation.w])
        self.velocity_data.append([current_time,latest_msg.angular_velocity.x,latest_msg.angular_velocity.y,latest_msg.angular_velocity.z])
        self.acceleration_data.append([current_time,latest_msg.linear_acceleration.x,latest_msg.linear_acceleration.y,latest_msg.linear_acceleration.z])
        #in the above 3 line of code, based on the type of the data (whether it is orientation, velocity or acceleration data), it is appended in the corresponding lists. 
        
        self.message_buffer.clear()


    
    def save_to_excel(self):
        # Append collected data to the sheet
        for a in self.orientation_data:
            self.orientation_worksheet.append(a)

        for b in self.velocity_data:
            self.velocity_worksheet.append(b)

        for c in self.acceleration_data:
            self.acceleration_worksheet.append(c)

        # Save the workbook
        self.workbook.save('imu.xlsx')
        self.get_logger().info('Data saved to imu_subscriber.xlsx')

    def destroy_node(self):
        
        # Save data to Excel before shutting down the node
        self.get_logger().info('Saving data to Excel...')
        self.save_to_excel()
        super().destroy_node()
        



def main(args=None):
    rclpy.init(args=args)
    imu_subscriber = ImuSubscriber()
    
    try:    
         rclpy.spin(imu_subscriber)
        # the imu_subscriber node will start receiving and processing the messages indefinitely in a loop, until it is stopped manualy 
    except KeyboardInterrupt:    #when the user press ctrl+c to stop the node
         
         pass
    finally: #finally makes sure that task written under it will be followed after try and exception blocks, regardless of whether an eexception was raised or not
        if imu_subscriber:  # this if checks whether imu_subscriber exist or not
            imu_subscriber.destroy_node() #function to destroy the node is called 
        if rclpy.ok(): #this checks if rclpy is still running. If it is running then rclpy.shutdown will be executed
            rclpy.shutdown()
     


if __name__ == '__main__':
    main()
